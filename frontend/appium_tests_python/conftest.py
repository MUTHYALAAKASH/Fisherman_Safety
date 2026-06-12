import pytest
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_string(val):
    if not isinstance(val, str):
        return val
    # Strip ANSI escape codes (color/formatting escapes)
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    val = ansi_escape.sub('', val)
    # Remove control characters with ASCII < 32 (excluding tab, lf, cr)
    return "".join(ch for ch in val if ord(ch) >= 32 or ch in "\t\n\r")

# List to gather all test execution statistics
test_results = []

def pytest_runtest_setup(item):
    name = item.name
    if name.startswith("test_func_"):
        section = "Functional Testing"
    elif name.startswith("test_ui_"):
        section = "UI-UX Testing"
    elif name.startswith("test_val_"):
        section = "Validation Testing"
    elif name.startswith("test_unit_"):
        section = "Unit Testing"
    else:
        section = "Uncategorized"
    
    print(f"\n[RUNNING] [Section: {section}] - Test: {name}")

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    # Execute the downstream hooks to obtain the actual report object
    outcome = yield
    rep = outcome.get_result()
    
    # We want to record the status when the test code is executed (the "call" phase)
    if rep.when == "call":
        # Extract function docstring as the description of the test case
        description = item.obj.__doc__.strip() if item.obj.__doc__ else "No description provided."
        
        # Determine category based on prefix
        name = item.name
        if name.startswith("test_func_"):
            category = "Functional Testing"
        elif name.startswith("test_ui_"):
            category = "UI-UX Testing"
        elif name.startswith("test_val_"):
            category = "Validation Testing"
        elif name.startswith("test_unit_"):
            category = "Unit Testing"
        else:
            category = "Uncategorized"

        # Capture failure traceback details if any
        error_msg = ""
        if rep.failed:
            error_msg = "Unknown error occurred"
            if rep.longrepr:
                if hasattr(rep.longrepr, 'reprcrash'):
                    error_msg = str(rep.longrepr.reprcrash.message)
                else:
                    error_msg = str(rep.longrepr)
            print(f" -> [FAILED] - Test: {name}")
            clean_err = clean_string(error_msg)
            print(f"    Reason for failure:\n{clean_err}")
        else:
            print(f" -> [PASSED] - Test: {name}")
        
        test_results.append({
            "Test Case": name,
            "Description": description,
            "Result": rep.outcome.upper(),
            "Duration (sec)": round(rep.duration, 4),
            "Error Message": error_msg,
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Category": category
        })

def pytest_sessionfinish(session, exitstatus):
    """
    Called after all tests have completed. Generates a beautifully formatted multi-sheet Excel report.
    """
    if not test_results:
        print("\n[Report] No test results were collected. Skipping Excel generation.")
        return

    # Create new workbook
    wb = openpyxl.Workbook()
    
    # Style definitions
    font_family = "Segoe UI"
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Color fills
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # soft green
    pass_font_color = "006100"
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # soft red
    fail_font_color = "9C0006"
    
    # ==========================================
    # Sheet 1: Summary Dashboard
    # ==========================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Banner Block
    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Smart Fisherman Safety - Automation Execution Dashboard"
    title_cell.font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    # Metadata Block
    ws_summary["A3"] = "Project Name:"
    ws_summary["B3"] = "Smart Fisherman Safety Mobile App"
    ws_summary["A4"] = "Generated At:"
    ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in ["3", "4"]:
        ws_summary["A" + row].font = Font(name=font_family, size=11, bold=True, color="595959")
        ws_summary["B" + row].font = Font(name=font_family, size=11)
        
    # Calculate Summary Statistics
    total_tests = len(test_results)
    passed_tests = sum(1 for r in test_results if r["Result"] == "PASSED")
    failed_tests = sum(1 for r in test_results if r["Result"] == "FAILED")
    pass_rate = round((passed_tests / total_tests) * 100, 2) if total_tests > 0 else 0.0
    deployable = "DEPLOYABLE" if failed_tests == 0 else "NOT DEPLOYABLE"
    
    # KPI Block (Side-by-side KPI cards)
    # Total Run Card
    ws_summary.merge_cells("A6:A7")
    kpi_run = ws_summary["A6"]
    kpi_run.value = f"Total Runs\n{total_tests}"
    kpi_run.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    kpi_run.font = Font(name=font_family, size=12, bold=True, color="1F4E78")
    kpi_run.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    kpi_run.border = thin_border
    
    # Passed Card
    ws_summary.merge_cells("B6:B7")
    kpi_pass = ws_summary["B6"]
    kpi_pass.value = f"Passed\n{passed_tests}"
    kpi_pass.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    kpi_pass.font = Font(name=font_family, size=12, bold=True, color=pass_font_color)
    kpi_pass.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    kpi_pass.border = thin_border
    
    # Failed Card
    ws_summary.merge_cells("C6:C7")
    kpi_fail = ws_summary["C6"]
    kpi_fail.value = f"Failed\n{failed_tests}"
    kpi_fail.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    kpi_fail.font = Font(name=font_family, size=12, bold=True, color=fail_font_color)
    kpi_fail.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    kpi_fail.border = thin_border
    
    # Pass Rate Card
    ws_summary.merge_cells("D6:D7")
    kpi_rate = ws_summary["D6"]
    kpi_rate.value = f"Pass Rate\n{pass_rate}%"
    kpi_rate.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    kpi_rate.font = Font(name=font_family, size=12, bold=True, color="305496")
    kpi_rate.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    kpi_rate.border = thin_border

    # Deployability Status Card
    ws_summary.merge_cells("E6:E7")
    kpi_dep = ws_summary["E6"]
    kpi_dep.value = f"Deployability\n{deployable}"
    kpi_dep.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    kpi_dep.font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    kpi_dep.fill = navy_fill if failed_tests == 0 else PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    kpi_dep.border = thin_border
    
    ws_summary.row_dimensions[6].height = 20
    ws_summary.row_dimensions[7].height = 20

    # Categorized Breakdown Section
    ws_summary["A9"] = "Test Category Breakdown"
    ws_summary["A9"].font = Font(name=font_family, size=14, bold=True, color="1F4E78")
    
    breakdown_headers = ["Test Category", "Total Run", "Passed", "Failed", "Pass Rate (%)"]
    ws_summary.append([]) # spacer row
    ws_summary.append(breakdown_headers)
    ws_summary.row_dimensions[11].height = 25
    
    for col_idx, h in enumerate(breakdown_headers, 1):
        cell = ws_summary.cell(row=11, column=col_idx)
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    categories = ["Functional Testing", "UI-UX Testing", "Validation Testing", "Unit Testing"]
    current_row = 12
    for cat in categories:
        cat_tests = [r for r in test_results if r["Category"] == cat]
        cat_total = len(cat_tests)
        cat_pass = sum(1 for r in cat_tests if r["Result"] == "PASSED")
        cat_fail = sum(1 for r in cat_tests if r["Result"] == "FAILED")
        cat_rate = round((cat_pass / cat_total) * 100, 2) if cat_total > 0 else 0.0
        
        ws_summary.append([cat, cat_total, cat_pass, cat_fail, f"{cat_rate}%"])
        
        for col_idx in range(1, 6):
            cell = ws_summary.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name=font_family, size=11)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        ws_summary.row_dimensions[current_row].height = 22
        current_row += 1
        
    # Set summary column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 20

    # ==========================================
    # Sheets 2-5: Detailed Test Records by Category
    # ==========================================
    sheets_config = {
        "Functional Testing": "Functional Tests",
        "UI-UX Testing": "UI-UX Tests",
        "Validation Testing": "Validation Tests",
        "Unit Testing": "Unit Tests"
    }
    
    headers = ["Test Case Name", "Description / Objectives", "Status", "Duration (sec)", "Failure Reason (if any)", "Timestamp"]
    
    for cat_name, sheet_title in sheets_config.items():
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        # Category Title Banner Block
        ws.merge_cells("A1:F1")
        banner = ws["A1"]
        banner.value = f"Detailed Test Log - {cat_name}"
        banner.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
        banner.fill = navy_fill
        banner.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35
        
        # Spacer & Headers
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Add test cases in this category
        cat_results = [r for r in test_results if r["Category"] == cat_name]
        
        current_row = 4
        for res in cat_results:
            ws.append([
                clean_string(res["Test Case"]),
                clean_string(res["Description"]),
                clean_string(res["Result"]),
                res["Duration (sec)"],
                clean_string(res["Error Message"]),
                clean_string(res["Timestamp"])
            ])
            
            # Format rows
            status_cell = ws.cell(row=current_row, column=3)
            if res["Result"] == "PASSED":
                status_cell.fill = pass_fill
                status_cell.font = Font(name=font_family, size=11, bold=True, color=pass_font_color)
            else:
                status_cell.fill = fail_fill
                status_cell.font = Font(name=font_family, size=11, bold=True, color=fail_font_color)
                
            for col_idx in range(1, 7):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                if col_idx != 3:
                    cell.font = Font(name=font_family, size=11)
                
                # Layout aligns
                if col_idx in [1, 2, 5]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx in [3, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 4:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            
        # Adjust Column Widths
        ws.column_dimensions['A'].width = 32 # Test Case Name
        ws.column_dimensions['B'].width = 50 # Description
        ws.column_dimensions['C'].width = 12 # Status
        ws.column_dimensions['D'].width = 15 # Duration
        ws.column_dimensions['E'].width = 45 # Error Message
        ws.column_dimensions['F'].width = 20 # Timestamp

    # Save Excel workbook
    report_name = "appium_test_report.xlsx"
    try:
        wb.save(report_name)
        print(f"\n[Excel Dashboard Reporter] Multi-sheet execution report saved to: {report_name}")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_name = f"appium_test_report_{timestamp}.xlsx"
        try:
            wb.save(fallback_name)
            print(f"\n[WARNING] Could not save to '{report_name}' because the file is open/locked in another program.")
            print(f"[Excel Dashboard Reporter] Saved fallback report to: {fallback_name}")
        except Exception as e:
            print(f"\n[ERROR] Failed to save Excel report: {e}")
    except Exception as e:
        print(f"\n[ERROR] Failed to save Excel report: {e}")
